'''
Created on Nov 25, 2021

@author: Fabian Ponce
'''
from datetime import datetime, date
import os, errno
import pyodbc
#import pandas as pd
from openpyxl import load_workbook
import openpyxl


#import csv
#import codecs


str_Line = '___________________________________________________'
str_Path = '.\OutPut\#_JIRA_Report.XLSX'

#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
def fn_B02_Write_Workbook_Body():
    print(str_Line, " fn_B02_Write_Workbook_Body")
    ###-----------------------------------------------------------------------------------------------------
    file_path = str_Path    
    wb = load_workbook(file_path)    
    ws = wb['JIRA STORIES']  # or wb.active    
    ###-----------------------------------------------------------------------------------------------------
    str_Query_01 = 'select  top 10 * from [JIRA].[dbo].[V_JIRA_STORIES]'    
    #print('str_Query_01 = ', str_Query_01)    
    conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=DESKTOP-C1P2GG2;'
                              'Database=JIRA;'
                              'Trusted_Connection=yes;')    
    cursor = conn.cursor()
    cursor.execute(str_Query_01)      
    i = 1    
    for row in cursor: 
        i = i +1
        #print(i , ' - ', row[0])
        print(i , ' - ', row[0] , ' - ', row[1], ' - ', row[2], ' - ', row[3], ' - ', row[4], ' - ', row[5], ' - ', row[6] , ' - ', row[7] , ' - ', row[8] , ' - ', row[9] , ' - ', row[10] , ' - ', row[11] , ' - ', row[12] )
        #print(row[0])  
 

        #ws.write(i, 0, row[0])
        ws['B5'] ='hola?'
    '''     
            V_Sheet_1.cell(row=i,column=1).value = row[0]
            V_Sheet_1.cell(row=i,column=2).value = row[1]
            V_Sheet_1.cell(row=i,column=3).value = row[2]
            V_Sheet_1.cell(row=i,column=4).value = row[3]
            V_Sheet_1.cell(row=i,column=5).value = row[4]
    '''        
    cursor.close()
    conn.close()
    ###-----------------------------------------------------------------------------------------------------
    wb.save(file_path)    
        
    ###-----------------------------------------------------------------------------------------------------    
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================

def fn_B01_Write_Workbook():
    print(str_Line, " fn_B01_Write_Workbook")
    #from openpyxl import load_workbook
    
    file_path = str_Path    
    wb = load_workbook(file_path)    
    ws = wb['JIRA STORIES']  # or wb.active
    
    #ws['G6'] = 123
    ws['A1'] ='[DT]'
    ws['B1'] ='[DT_SPRINT]'
    ws['C1'] ='[SPRINT]'
    ws['D1'] ='[Issue_Type]'
    ws['E1'] ='[Summary]'
    ws['F1'] ='[Key]'
    ws['G1'] ='[Assignee]'
    ws['H1'] ='[Reporter]'
    ws['I1'] ='[Priority]'
    ws['J1'] ='[Status]'
    ws['K1'] ='[Status2]'
    ws['L1'] ='[Resolution]'
    ws['M1'] ='[Created]'
    ws['N1'] ='[Updated]'
    ws['O1'] ='[Due_date]'
    ws['P1'] ='[parent]'
    ws['Q1'] ='[Resolved]'
    ws['R1'] ='[Status_Category_Changed]'
    ws['S1'] ='[Linked_Issues]'
    ws['T1'] ='[Parent_Link]'
    ws['U1'] ='[Time Resolution]'
    ws['V1'] ='[Wk_Created]'
    ws['W1'] ='[Wk_Resolved]'
    ws['X1'] ='[Wk_Created_Resolved]'
       
    wb.save(file_path)
    
    



#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================

def fn_A03_DELETE_File():
    print(str_Line, " fn_03_DELETE_File")
    
    if os.path.exists(str_Path):
      os.remove(str_Path)
      fn_A02_Crea_XLS()
    else:
      print("The file does not exist")    

def fn_A02_Crea_XLS():
    print(str_Line, " fn_02_Crea_XLS")
    
    wb = openpyxl.Workbook()
    hoja = wb.active
    
    hoja1 = wb.create_sheet("JIRA STORIES", 0)
    hoja2 = wb.create_sheet("RESOLVED_LW", 1)
    wb.create_sheet(index=2, title="JIRA_Open_Past_Q")
    print(wb.sheetnames)
    wb.save(str_Path)
    
    fn_B01_Write_Workbook()
    


def fn_A01_Crea_XLS():
    print(str_Line, " fn_01_Crea_XLS")
    #fileName = r"C:\Test\test.txt"
    fileName = str_Path
    if_FN_Exists = os.path.exists(fileName)
    #Out: True
    print('if_FN_Exists = ', if_FN_Exists)

    #Out: True    
    if if_FN_Exists == True:
        print('es True :)')
        fn_A03_DELETE_File()
    else:
        print('False  :)')
        fn_A02_Crea_XLS()
        

#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================

def main():
    print(str_Line, " main")
    #----------------------------------------------
    dt_now1 = datetime.now()
    #----------------------------------------------
    #----------------------------------------------
    fn_A01_Crea_XLS()



    print('\n\n\t------ Done :) \n\n')
    #----------------------------------------------
    #----------------------------------------------
    dt_now2 = datetime.now()
    print('---------START =', dt_now1)
    print('-----------END =', dt_now2)
    duration = dt_now2 - dt_now1                         # For build-in functions
    duration_in_s = duration.total_seconds()
    print('----------Time =', duration_in_s)
    #----------------------------------------------

if __name__ == "__main__":
    main()