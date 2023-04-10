# -*- coding: utf-8 -*-
'''
Created on Nov 25, 2021

@author: Fabian Ponce
'''
from datetime import datetime, date
import os, errno
import pyodbc
#import pandas as pd
from openpyxl import load_workbook
import openpyxl


#import csv
#import codecs

#####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
### https://zetcode.com/python/openpyxl/
#####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
str_Line = '___________________________________________________'
str_Path = '.\OutPut\#_JIRA_Report.XLSX'

#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
def fn_01_XLS_SHEET_JIRA_STORIES():
    print(str_Line, " fn_01_XLS_SHEET_JIRA_STORIES")
    ###-----------------------------------------------------------------------------------------------------
    file_path = str_Path    
    wb = load_workbook(file_path)    
    ws = wb['JIRA STORIES']  # or wb.active    
    ###-----------------------------------------------------------------------------------------------------
    ###-----------------------------------------------------------------------------------------------------
    ###-----------------------------------------------------------------------------------------------------
    str_Query_01 = 'select   * from [JIRA].[dbo].[V_JIRA_STORIES] order by DT'    
    #print('str_Query_01 = ', str_Query_01)    
    '''
    conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=MDC-DCMW8C3;'
                              'Database=JIRA;'
                              'Trusted_Connection=yes;')
                              '''
    server = r"MDC-DCMW8C3"
    
    db = "JIRA"
    user = "sa"
    password = "123456"
        
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server + ';DATABASE=' + db +';UID=' + user + ';PWD=' + password)
        
    cursor = conn.cursor()
    cursor.execute(str_Query_01)      
    i = 1    
    for row in cursor: 
        i = i +1
        print(i , ' - ', row[0])
        #print(i , ' - ', row[0] , ' - ', row[1], ' - ', row[2], ' - ', row[3], ' - ', row[4], ' - ', row[5], ' - ', row[6] , ' - ', row[7] , ' - ', row[8] , ' - ', row[9] , ' - ', row[10] , ' - ', row[11] , ' - ', row[12] )
        #print(row[0])  
 
        ###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
        ##ws.write(i, 0, row[0])
        #ws['B5'] ='hola?'
        #ws['B6'] =row[0]
        #ws.cell(row=2, column=2).value = row[0]

        
        ws.cell(row=i, column=1).value = row[0]
        ws.cell(row=i, column=2).value = row[1]
        ws.cell(row=i, column=3).value = row[2]
        ws.cell(row=i, column=4).value = row[3]
        ws.cell(row=i, column=5).value = row[4]     #[Summary]
        ws.cell(row=i, column=6).value = row[5]
        ws.cell(row=i, column=7).value = row[6]     #[Key]
        ws.cell(row=i, column=8).value = row[7]
        ws.cell(row=i, column=9).value = row[8]
        ws.cell(row=i, column=10).value = row[9]    #[Priority]
        ws.cell(row=i, column=11).value = row[10]
        ws.cell(row=i, column=12).value = row[11]
        ws.cell(row=i, column=13).value = row[12]
        ws.cell(row=i, column=14).value = row[13]   #[Created]
        ws.cell(row=i, column=15).value = row[14]
        ws.cell(row=i, column=16).value = row[15]
        ws.cell(row=i, column=17).value = row[16]   #[parent]
        ws.cell(row=i, column=18).value = row[17]
        ws.cell(row=i, column=19).value = row[18]   #[Status_Category_Changed]
        ws.cell(row=i, column=20).value = row[19]
        ws.cell(row=i, column=21).value = row[20]   #[Wk_Created]
        ws.cell(row=i, column=22).value = row[21]
        ws.cell(row=i, column=23).value = row[22]    
        #ws.cell(row=i, column=24).value = row[23]
        
        ###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%

    cursor.close()
    conn.close()
    ###-----------------------------------------------------------------------------------------------------        
    ###-----------------------------------------------------------------------------------------------------     



    ###-----------------------------------------------------------------------------------------------------
    ###-----------------------------------------------------------------------------------------------------
    wb.save(file_path)            
    ###-----------------------------------------------------------------------------------------------------        
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
def fn_B02_Write_Workbook_Body():
    print(str_Line, " fn_B02_Write_Workbook_Body")
    ###-----------------------------------------------------------------------------------------------------
    file_path = str_Path    
    wb = load_workbook(file_path)    
    ws = wb['JIRA STORIES']  # or wb.active    
    ###-----------------------------------------------------------------------------------------------------


        #ws.write(i, 0, row[0])
    ws['B5'] ='hola?'

    ###-----------------------------------------------------------------------------------------------------
    wb.save(file_path)    
        
    ###-----------------------------------------------------------------------------------------------------    
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
def fn_B02_Write_Workbook_Body_SQL():
    print(str_Line, " fn_B02_Write_Workbook_Body_SQL")
    ###-----------------------------------------------------------------------------------------------------
    ###-----------------------------------------------------------------------------------------------------
    str_Query_01 = 'select   * from [JIRA].[dbo].[V_JIRA_STORIES] order by DT'    
    #print('str_Query_01 = ', str_Query_01)    
    conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=DESKTOP-C1P2GG2;'
                              'Database=JIRA;'
                              'Trusted_Connection=yes;')    
    cursor = conn.cursor()
    cursor.execute(str_Query_01)      
    i = 1    
    for row in cursor: 
        i = i +1
        #print(i , ' - ', row[0])
        print(i , ' - ', row[0] , ' - ', row[1], ' - ', row[2], ' - ', row[3], ' - ', row[4], ' - ', row[5], ' - ', row[6] , ' - ', row[7] , ' - ', row[8] , ' - ', row[9] , ' - ', row[10] , ' - ', row[11] , ' - ', row[12] )
        print(row[0])  
 


    '''     
            V_Sheet_1.cell(row=i,column=1).value = row[0]
            V_Sheet_1.cell(row=i,column=2).value = row[1]
            V_Sheet_1.cell(row=i,column=3).value = row[2]
            V_Sheet_1.cell(row=i,column=4).value = row[3]
            V_Sheet_1.cell(row=i,column=5).value = row[4]
    '''        
    cursor.close()
    conn.close()
    ###-----------------------------------------------------------------------------------------------------        
    ###-----------------------------------------------------------------------------------------------------        

        

#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================

def main():
    print(str_Line, " main")
    #----------------------------------------------
    dt_now1 = datetime.now()
    #----------------------------------------------
    #----------------------------------------------
    #fn_B02_Write_Workbook_Body()
    #fn_B02_Write_Workbook_Body_SQL()
    
    fn_01_XLS_SHEET_JIRA_STORIES()



    print('\n\n\t------ Done :) \n\n')
    #----------------------------------------------
    #----------------------------------------------
    dt_now2 = datetime.now()
    print('---------START =', dt_now1)
    print('-----------END =', dt_now2)
    duration = dt_now2 - dt_now1                         # For build-in functions
    duration_in_s = duration.total_seconds()
    print('----------Time =', duration_in_s)
    #----------------------------------------------

if __name__ == "__main__":
    main()