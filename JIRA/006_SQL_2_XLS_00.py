'''
Created on Nov 25, 2021

@author: Fabian Ponce
'''
from datetime import datetime, date
import os, errno
import pyodbc
#import pandas as pd
from openpyxl import load_workbook
import openpyxl


#import csv
#import codecs


str_Line = '___________________________________________________'
str_Path = '.\OutPut\#_JIRA_Report.XLSX'

#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####=============================================================================================== 
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
def fn_B03_Write_Workbook():
    print(str_Line, " fn_B03_Write_Workbook")
    
    file_path = str_Path    
    wb = load_workbook(file_path)    
    ws = wb['JIRA_Open_Past_Q']  # or wb.active
    
    ws['B1'] ='[DT]'
    ws['C1'] ='[DT_SPRINT]'
    ws['D1'] ='[SPRINT]'
    ws['E1'] ='[Issue_Type]'
    ws['F1'] ='[Summary]'
    ws['G1'] ='[Key]'
    ws['H1'] ='[Assignee]'
    ws['I1'] ='[Reporter]'
    ws['J1'] ='[Priority]'
    ws['K1'] ='[Status]'
    ws['L1'] ='[Resolution]'
    ws['M1'] ='[Created]'
    ws['N1'] ='[QUARTER]'
    ws['O1'] ='[Updated]'
    ws['P1'] ='[Due_date]'
    ws['Q1'] ='[parent]'
    ws['R1'] ='[Resolved]'
    ws['S1'] ='[Status_Category_Changed]'
    #ws['S1'] ='[Linked_Issues]'
    #ws['T1'] ='[Parent_Link]'
    #ws['U1'] ='[Time Resolution]'
    
    ws['T1'] ='[Time Resolution]'
       
    wb.save(file_path)
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
def fn_B02_Write_Workbook():
    print(str_Line, " fn_B02_Write_Workbook")
    
    file_path = str_Path    
    wb = load_workbook(file_path)    
    ws = wb['RESOLVED_LW']  # or wb.active
    
    ws['A1'] ='[DT]'
    ws['B1'] ='[DT_SPRINT]'
    ws['C1'] ='[SPRINT]'
    ws['D1'] ='[Issue_Type]'
    ws['E1'] ='[Summary]'
    ws['F1'] ='[Key]'
    ws['G1'] ='[Assignee]'
    ws['H1'] ='[Reporter]'
    ws['I1'] ='[Priority]'
    ws['J1'] ='[Status]'
    ws['K1'] ='[Resolution]'
    ws['L1'] ='[Created]'
    ws['M1'] ='[Updated]'
    ws['N1'] ='[Due_date]'
    ws['O1'] ='[parent]'
    ws['P1'] ='[Resolved]'
    ws['Q1'] ='[Status_Category_Changed]'
    #ws['R1'] ='[Linked_Issues]'
    #ws['S1'] ='[Parent_Link]'
    #ws['T1'] ='[Wk_Created]'
    #ws['U1'] ='[Wk_Resolved]'
    #ws['V1'] ='[Wk_Created_Resolved]'
    
    ws['R1'] ='[Wk_Created]'
    ws['S1'] ='[Wk_Resolved]'
    ws['T1'] ='[Wk_Created_Resolved]'    
       
    wb.save(file_path)
#####===============================================================================================
#####===============================================================================================
def fn_B01_Write_Workbook():
    print(str_Line, " fn_B01_Write_Workbook")
    
    file_path = str_Path    
    wb = load_workbook(file_path)    
    ws = wb['JIRA STORIES']  # or wb.active
    
    ws['B1'] ='[DT]'
    ws['C1'] ='[DT_SPRINT]'
    ws['D1'] ='[SPRINT]'
    ws['E1'] ='[Issue_Type]'
    ws['F1'] ='[Summary]'
    ws['G1'] ='[Key]'
    ws['H1'] ='[Assignee]'
    ws['I1'] ='[Reporter]'
    ws['J1'] ='[Priority]'
    ws['K1'] ='[Status]'
    ws['L1'] ='[Status2]'
    ws['M1'] ='[Resolution]'
    ws['N1'] ='[Created]'
    ws['O1'] ='[Updated]'
    ws['P1'] ='[Due_date]'
    ws['Q1'] ='[parent]'
    ws['R1'] ='[Resolved]'
    ws['S1'] ='[Status_Category_Changed]'
    #ws['S1'] ='[Linked_Issues]'
    #ws['T1'] ='[Parent_Link]'
    #ws['U1'] ='[Time Resolution]'
    #ws['V1'] ='[Wk_Created]'
    #ws['W1'] ='[Wk_Resolved]'
    #ws['X1'] ='[Wk_Created_Resolved]'
    
    ws['T1'] ='[Time Resolution]'
    ws['U1'] ='[Wk_Created]'
    ws['V1'] ='[Wk_Resolved]'
    ws['W1'] ='[Wk_Created_Resolved]'
       
    wb.save(file_path)
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================

def fn_A03_DELETE_File():
    print(str_Line, " fn_03_DELETE_File")
    
    if os.path.exists(str_Path):
      os.remove(str_Path)
      fn_A02_Crea_XLS()
    else:
      print("The file does not exist")    

def fn_A02_Crea_XLS():
    print(str_Line, " fn_02_Crea_XLS")
    
    wb = openpyxl.Workbook()
    hoja = wb.active
    
    hoja1 = wb.create_sheet("JIRA STORIES", 0)
    hoja2 = wb.create_sheet("RESOLVED_LW", 1)
    wb.create_sheet(index=2, title="JIRA_Open_Past_Q")
    print(wb.sheetnames)
    wb.save(str_Path)
    
    fn_B01_Write_Workbook()
    fn_B02_Write_Workbook()
    fn_B03_Write_Workbook()


def fn_A01_Crea_XLS():
    print(str_Line, " fn_01_Crea_XLS")
    #fileName = r"C:\Test\test.txt"
    fileName = str_Path
    if_FN_Exists = os.path.exists(fileName)
    #Out: True
    print('if_FN_Exists = ', if_FN_Exists)

    #Out: True    
    if if_FN_Exists == True:
        print('es True :)')
        fn_A03_DELETE_File()
    else:
        print('False  :)')
        fn_A02_Crea_XLS()
        

#####===============================================================================================
#####===============================================================================================
#####===============================================================================================
#####===============================================================================================

def main():
    print(str_Line, " main")
    #----------------------------------------------
    dt_now1 = datetime.now()
    #----------------------------------------------
    #----------------------------------------------
    fn_A01_Crea_XLS()



    print('\n\n\t------ Done :) \n\n')
    #----------------------------------------------
    #----------------------------------------------
    dt_now2 = datetime.now()
    print('---------START =', dt_now1)
    print('-----------END =', dt_now2)
    duration = dt_now2 - dt_now1                         # For build-in functions
    duration_in_s = duration.total_seconds()
    print('----------Time =', duration_in_s)
    #----------------------------------------------

if __name__ == "__main__":
    main()