from datetime import datetime, date
import os, errno
import pandas as pd
from openpyxl import Workbook, load_workbook
import csv
import codecs

def fn_6_csv_from_excel(str_FileName,str_sheetnames):

    print("----------------fn_6_csv_from_excel---------------------------------")
    print('--- str_FileName = ', str_FileName)
    print('--- str_sheetnames = ' , str_sheetnames)  
    Len_str_FileName = 0
    Len_str_FileName = len(str_FileName)
    print('Len_str_FileName = ',  Len_str_FileName)
    str_FileName_A = str_FileName[0:Len_str_FileName-5]
    str_FileName_A = str_FileName_A+'_'+str_sheetnames+'.CSV'
    print('\nstr_FileName_A = ', str_FileName_A)
    str_FileName_A = str_FileName_A.replace("InPut","OutPut")
    print('str_FileName_A', str_FileName_A)               
    data_xls = pd.read_excel(str_FileName, sheet_name=str_sheetnames, index_col=None)      
    data_xls.to_csv(str_FileName_A, encoding='utf-8')    
    print("********************************************************************")


def fn_5_Xls_Sheets(str_FileName):
    print("----------------fn_5_Xls_Sheets---------------------------------")
    print('--- str_FileName = ', str_FileName)
    sheet_Name = load_workbook(str_FileName)
    str_sheetnames = sheet_Name.sheetnames
    print(str_sheetnames)
    print('---sheet_Name.sheetnames =', sheet_Name.sheetnames)
    # -- ['R2019_Oct', 'R2019_Nov', 'R2019_Dic', 'R2020_Ene', 'R2020_Feb', 'R2020_Mar', 'R2020_Abr']               
    sheet_Name.close()  
   
    for str_xxx in str_sheetnames:
        print('*******', str_xxx)
        fn_6_csv_from_excel(str_FileName,str_xxx)



def fn_4_FileList(list_allFiles):
    print("---------------fn_FileList-------------------------------------")
    print(list_allFiles)
 
    for str_FileName in list_allFiles:
        print(str_FileName)
        N_str_FileName_Len = len(str_FileName)
        print('N_str_FileName_Len = ', N_str_FileName_Len)
        str_A = str_FileName[1:N_str_FileName_Len]
        N_str_A_Len = len(str_A)
        #print('N_str_A_Len = ', N_str_A_Len)
        #print('str_A = ', str_A)
        N_str_A_ext = str_A.index('.')
        #print('N_str_A_ext = ', N_str_A_ext)   
        str_Ext = str_A[N_str_A_ext+1:N_str_A_Len]
        str_Ext = str_Ext.lower()
        #print('--------- str_Ext = ', str_Ext)      
        if str_Ext == 'xlsx':
            print('es un xlsx')
            fn_5_Xls_Sheets(str_FileName)
        elif str_Ext == 'csv':
            print('es un csv')
            #fn_CSV(str_FileName)
        else:
            print('--- else')

        #N_XLSX = str_FileName.find(".xlsx")      
        #print('N_XLSX = ', N_XLSX)
        #fn_Xls_Sheets(str_FileName)

def fn_3_getListOfFiles(dirName):
    print("---------------getListOfFiles-------------------------------------")
    # create a list of file and sub directories
    # names in the given directory
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory
        if os.path.isdir(fullPath):
            allFiles = allFiles + fn_3_getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)
    print(allFiles)
    print(allFiles[1])
    fn_4_FileList(allFiles)
def fn_2_Chk_Fldr_Item():
    print("---------------fn_Chk_Fldr_Item-------------------------------------")
    dirName = '.\InPut'
    if os.path.exists(dirName) and os.path.isdir(dirName):
        if not os.listdir(dirName):
            print("Directory is empty !!!")
        else:  
            print("Directory is not empty")
            fn_3_getListOfFiles(dirName)
    else:
        print("Given Directory don't exists")
def fn_1_Chk_Fldr_InOutPut():
    print("---------------fn_Chk_Fldr_InOutPut----------------------------------")
    path = '.\InPut'
    try:
        os.makedirs(path)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise

    path = '.\OutPut'
    try:
        os.makedirs(path)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise 

def main():
    print("-------------------------------------------------")
    #----------------------------------------------
    dt_now1 = datetime.now()
    #----------------------------------------------
    #----------------------------------------------
    fn_1_Chk_Fldr_InOutPut()
    fn_2_Chk_Fldr_Item()
    #fn_Xls_Sheets()
    #fn_csv_from_excel()
    print('\n\n\t------ Done :) \n\n')
    #----------------------------------------------
    #----------------------------------------------
    dt_now2 = datetime.now()
    print('---------START =', dt_now1)
    print('-----------END =', dt_now2)
    duration = dt_now2 - dt_now1                         # For build-in functions
    duration_in_s = duration.total_seconds()
    print('----------Time =', duration_in_s)
    #----------------------------------------------

if __name__ == "__main__":
    main()